#!/usr/bin/env python3
"""
Export to Excel - Gerador de Relatórios Excel Formatados

Gera relatório Excel com formatação profissional a partir do CSV de análise.
"""

import sys
import csv
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelReportGenerator:
    """Gerador de relatórios Excel"""

    def __init__(self, csv_path: str, output_path: str):
        self.csv_path = Path(csv_path)
        self.output_path = Path(output_path)
        self.data = []
        self.summary_stats = {}

    def load_csv(self):
        """Carrega dados do CSV"""
        with open(self.csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            self.data = list(reader)

        # Calcular estatísticas
        self.calculate_summary()

    def calculate_summary(self):
        """Calcula estatísticas do relatório"""
        if not self.data:
            return

        total = len(self.data)
        conforme = sum(1 for row in self.data if row.get('veredicto', '').upper() == 'CONFORME')
        nao_conforme = sum(1 for row in self.data if row.get('veredicto', '').upper() == 'NAO_CONFORME')
        revisao = sum(1 for row in self.data if row.get('veredicto', '').upper() == 'REVISAO')

        # Calcular média de confiança
        confidences = []
        for row in self.data:
            try:
                conf = float(row.get('confianca', 0))
                confidences.append(conf)
            except (ValueError, TypeError):
                pass

        avg_confidence = sum(confidences) / len(confidences) if confidences else 0

        self.summary_stats = {
            'total': total,
            'conforme': conforme,
            'nao_conforme': nao_conforme,
            'revisao': revisao,
            'conforme_pct': (conforme / total * 100) if total > 0 else 0,
            'nao_conforme_pct': (nao_conforme / total * 100) if total > 0 else 0,
            'revisao_pct': (revisao / total * 100) if total > 0 else 0,
            'avg_confidence': avg_confidence,
        }

    def create_summary_sheet(self, wb: Workbook):
        """Cria aba de resumo"""
        ws = wb.create_sheet("Resumo", 0)

        # Estilos
        title_font = Font(name='Arial', size=16, bold=True, color='1F4788')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws['A1'] = 'RELATÓRIO DE ANÁLISE DE EDITAL'
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 25

        # Informações
        ws['A3'] = 'Edital:'
        ws['B3'] = self.csv_path.parent.name
        ws['A4'] = 'Data de Análise:'
        ws['B4'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        ws['A5'] = 'Sistema:'
        ws['B5'] = 'BidAnalyzee v2.0'

        # Resumo Executivo
        ws['A7'] = 'RESUMO EXECUTIVO'
        ws['A7'].font = Font(name='Arial', size=14, bold=True, color='1F4788')
        ws.merge_cells('A7:D7')

        # Cabeçalho da tabela
        headers = ['Métrica', 'Quantidade', 'Percentual', 'Status']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=8, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Dados da tabela
        data_rows = [
            ['Total de Requisitos', self.summary_stats['total'], '100%', '📊'],
            ['Conforme', self.summary_stats['conforme'],
             f"{self.summary_stats['conforme_pct']:.1f}%", '✅'],
            ['Não Conforme', self.summary_stats['nao_conforme'],
             f"{self.summary_stats['nao_conforme_pct']:.1f}%", '❌'],
            ['Requer Revisão', self.summary_stats['revisao'],
             f"{self.summary_stats['revisao_pct']:.1f}%", '⚠️'],
        ]

        for row_idx, row_data in enumerate(data_rows, start=9):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Cores por linha
                if row_idx == 10:  # Conforme
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif row_idx == 11:  # Não Conforme
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                elif row_idx == 12:  # Revisão
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

        # Estatísticas adicionais
        ws['A14'] = 'Confiança Média:'
        ws['B14'] = f"{self.summary_stats['avg_confidence']:.2f}"
        ws['A14'].font = Font(bold=True)

        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10

        # Adicionar gráfico
        self.add_chart(ws)

    def add_chart(self, ws):
        """Adiciona gráfico de pizza à aba de resumo"""
        try:
            # Gráfico de pizza
            pie = PieChart()
            labels = Reference(ws, min_col=1, min_row=10, max_row=12)
            data = Reference(ws, min_col=2, min_row=9, max_row=12)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Distribuição de Vereditos"

            ws.add_chart(pie, "F8")
        except Exception as e:
            print(f"⚠️  Aviso: Não foi possível adicionar gráfico: {e}")

    def create_details_sheet(self, wb: Workbook):
        """Cria aba de detalhes"""
        ws = wb.create_sheet("Análise Detalhada", 1)

        # Estilos
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Cabeçalhos
        headers = ['ID', 'Descrição', 'Categoria', 'Criticidade',
                  'Veredicto', 'Confiança', 'Justificativa', 'Recomendações']

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Dados
        for row_idx, item in enumerate(self.data, start=2):
            ws.cell(row=row_idx, column=1).value = item.get('id', '')
            ws.cell(row=row_idx, column=2).value = item.get('descricao', '')
            ws.cell(row=row_idx, column=3).value = item.get('categoria', '')
            ws.cell(row=row_idx, column=4).value = item.get('criticidade', '')
            ws.cell(row=row_idx, column=5).value = item.get('veredicto', '')
            ws.cell(row=row_idx, column=6).value = item.get('confianca', '')
            ws.cell(row=row_idx, column=7).value = item.get('justificativa', '')
            ws.cell(row=row_idx, column=8).value = item.get('recomendacoes', '')

            # Aplicar borda
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).border = border
                ws.cell(row=row_idx, column=col).alignment = Alignment(
                    horizontal='left',
                    vertical='top',
                    wrap_text=True
                )

            # Cores por veredicto
            veredicto = item.get('veredicto', '').upper()
            if veredicto == 'CONFORME':
                fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif veredicto == 'NAO_CONFORME':
                fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif veredicto == 'REVISAO':
                fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            else:
                fill = None

            if fill:
                for col in range(1, 9):
                    ws.cell(row=row_idx, column=col).fill = fill

        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 50
        ws.column_dimensions['H'].width = 50

        # Congelar primeira linha
        ws.freeze_panes = 'A2'

    def create_filtered_sheets(self, wb: Workbook):
        """Cria abas filtradas por veredicto"""
        for veredicto, sheet_name in [
            ('NAO_CONFORME', 'Não Conformes'),
            ('REVISAO', 'Requer Revisão'),
            ('CONFORME', 'Conformes')
        ]:
            filtered_data = [row for row in self.data
                           if row.get('veredicto', '').upper() == veredicto]

            if not filtered_data:
                continue

            ws = wb.create_sheet(sheet_name)

            # Cabeçalhos
            headers = ['ID', 'Descrição', 'Justificativa', 'Recomendações']
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Dados
            for row_idx, item in enumerate(filtered_data, start=2):
                ws.cell(row=row_idx, column=1).value = item.get('id', '')
                ws.cell(row=row_idx, column=2).value = item.get('descricao', '')
                ws.cell(row=row_idx, column=3).value = item.get('justificativa', '')
                ws.cell(row=row_idx, column=4).value = item.get('recomendacoes', '')

                for col in range(1, 5):
                    ws.cell(row=row_idx, column=col).alignment = Alignment(
                        wrap_text=True,
                        vertical='top'
                    )

            # Ajustar largura
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 50

    def generate(self) -> bool:
        """Gera o Excel"""
        if not OPENPYXL_AVAILABLE:
            print("❌ openpyxl não está instalado")
            print("   Instale com: pip install openpyxl")
            return False

        try:
            # Carregar dados
            self.load_csv()

            # Criar workbook
            wb = Workbook()

            # Remover aba padrão
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Criar abas
            self.create_summary_sheet(wb)
            self.create_details_sheet(wb)
            self.create_filtered_sheets(wb)

            # Salvar
            wb.save(str(self.output_path))

            return True

        except Exception as e:
            print(f"❌ Erro ao gerar Excel: {e}")
            import traceback
            traceback.print_exc()
            return False


def main():
    """Entry point"""
    if len(sys.argv) < 2:
        print("\n❌ Uso incorreto\n")
        print("Uso: python3 scripts/export_excel.py <csv_path> [output_path]")
        print("\nExemplo:")
        print("   python3 scripts/export_excel.py data/.../analysis_results.csv")
        print("   python3 scripts/export_excel.py data/.../analysis_results.csv report.xlsx")
        print()
        sys.exit(1)

    csv_path = sys.argv[1]

    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        # Gerar nome automático
        csv_p = Path(csv_path)
        output_path = csv_p.parent / f"{csv_p.stem}_report.xlsx"

    print(f"\n📊 Gerando relatório Excel...")
    print(f"   CSV: {csv_path}")
    print(f"   Output: {output_path}")

    generator = ExcelReportGenerator(csv_path, output_path)
    success = generator.generate()

    if success:
        print(f"\n✅ Relatório Excel gerado com sucesso!")
        print(f"   📁 {output_path}")
        print(f"\n📋 Abas criadas:")
        print(f"   - Resumo (estatísticas e gráfico)")
        print(f"   - Análise Detalhada (todos os requisitos)")
        print(f"   - Não Conformes")
        print(f"   - Requer Revisão")
        print(f"   - Conformes")
    else:
        print(f"\n❌ Falha ao gerar relatório Excel")
        sys.exit(1)


if __name__ == "__main__":
    main()
