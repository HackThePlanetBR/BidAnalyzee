"""
Conformity Report Generator

Generates consolidated reports combining Document Structurer extraction
with Technical Analyst conformity analysis.

This module provides:
- ConformityReport: Data structure for consolidated analysis
- ReportExporter: Multi-format export (JSON, CSV, Excel, Markdown)
"""

from dataclasses import dataclass, field
from typing import Dict, List, Any
from pathlib import Path
import json
import csv


@dataclass
class ConformityReport:
    """
    Consolidated conformity report

    Combines edital metadata, extracted requirements, and conformity analysis
    into a single unified report structure.

    Attributes:
        edital_metadata: Metadata extracted from edital (Document Structurer)
        requirements: List of extracted requirements
        analysis_results: List of ConformityAnalysis objects (Query Processor)
        timestamp: ISO timestamp of analysis
        pipeline_stats: Performance statistics from pipeline execution
    """

    edital_metadata: Dict[str, Any]
    requirements: List[Dict[str, Any]]
    analysis_results: List[Any]  # List[ConformityAnalysis]
    timestamp: str
    pipeline_stats: Dict[str, Any] = field(default_factory=dict)

    def get_summary(self) -> Dict[str, Any]:
        """
        Get high-level summary statistics

        Returns:
            Dictionary with summary statistics including:
            - total_requirements: Total count
            - conforme, nao_conforme, revisao: Counts by verdict
            - Percentages for each category
        """
        total = len(self.analysis_results)
        conforme = sum(
            1 for r in self.analysis_results
            if r.conformity.value == 'CONFORME'
        )
        nao_conforme = sum(
            1 for r in self.analysis_results
            if r.conformity.value == 'NAO_CONFORME'
        )
        revisao = sum(
            1 for r in self.analysis_results
            if r.conformity.value == 'REVISAO'
        )

        return {
            'total_requirements': total,
            'conforme': conforme,
            'nao_conforme': nao_conforme,
            'revisao': revisao,
            'conforme_pct': (conforme / total * 100) if total > 0 else 0,
            'nao_conforme_pct': (nao_conforme / total * 100) if total > 0 else 0,
            'revisao_pct': (revisao / total * 100) if total > 0 else 0,
            'overall_compliance_rate': (conforme / total * 100) if total > 0 else 0
        }

    def get_recommendations(self) -> List[str]:
        """
        Get consolidated recommendations from all analyses

        Returns:
            List of unique recommendations across all requirements
        """
        all_recommendations = []
        for analysis in self.analysis_results:
            all_recommendations.extend(analysis.recommendations)

        # Return unique recommendations
        return list(set(all_recommendations))

    def get_critical_issues(self) -> List[Dict[str, Any]]:
        """
        Get list of non-conformant requirements

        Returns:
            List of requirements marked as NAO_CONFORME
        """
        issues = []
        for req, analysis in zip(self.requirements, self.analysis_results):
            if analysis.conformity.value == 'NAO_CONFORME':
                issues.append({
                    'requirement': req,
                    'analysis': analysis.to_dict()
                })
        return issues

    def get_review_needed(self) -> List[Dict[str, Any]]:
        """
        Get list of requirements needing manual review

        Returns:
            List of requirements marked as REVISAO
        """
        review = []
        for req, analysis in zip(self.requirements, self.analysis_results):
            if analysis.conformity.value == 'REVISAO':
                review.append({
                    'requirement': req,
                    'analysis': analysis.to_dict()
                })
        return review

    def to_dict(self) -> Dict[str, Any]:
        """
        Convert report to dictionary

        Returns:
            Complete report as nested dictionary
        """
        return {
            'edital_metadata': self.edital_metadata,
            'summary': self.get_summary(),
            'requirements': self.requirements,
            'analysis_results': [r.to_dict() for r in self.analysis_results],
            'critical_issues': self.get_critical_issues(),
            'review_needed': self.get_review_needed(),
            'consolidated_recommendations': self.get_recommendations(),
            'timestamp': self.timestamp,
            'pipeline_stats': self.pipeline_stats
        }

    def to_json(self, indent: int = 2) -> str:
        """
        Convert to JSON string

        Args:
            indent: JSON indentation level

        Returns:
            JSON string representation
        """
        return json.dumps(self.to_dict(), indent=indent, ensure_ascii=False)


class ReportExporter:
    """
    Export conformity reports in multiple formats

    Supports:
    - JSON: Complete structured data
    - CSV: Tabular format for spreadsheets
    - Excel: Multi-sheet workbook with summary and details
    - Markdown: Human-readable report

    Example:
        >>> exporter = ReportExporter(report, output_dir="output")
        >>> exporter.export("edital_001", "json")
        Path('output/edital_001_analysis.json')
    """

    def __init__(self, report: ConformityReport, output_dir: Path):
        """
        Initialize exporter

        Args:
            report: ConformityReport to export
            output_dir: Directory for output files
        """
        self.report = report
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export(self, basename: str, format: str) -> Path:
        """
        Export report in specified format

        Args:
            basename: Base filename (without extension)
            format: Export format ('json', 'csv', 'excel', 'markdown')

        Returns:
            Path to exported file

        Raises:
            ValueError: If format is not supported
        """
        if format == 'json':
            return self.to_json(basename)
        elif format == 'csv':
            return self.to_csv(basename)
        elif format == 'excel':
            return self.to_excel(basename)
        elif format == 'markdown':
            return self.to_markdown(basename)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def to_json(self, basename: str) -> Path:
        """
        Export as JSON

        Args:
            basename: Base filename

        Returns:
            Path to JSON file
        """
        filepath = self.output_dir / f"{basename}_analysis.json"
        filepath.write_text(self.report.to_json(), encoding='utf-8')
        return filepath

    def to_csv(self, basename: str) -> Path:
        """
        Export as CSV (requirements + analysis)

        Creates a CSV with one row per requirement, including:
        - Requirement fields (id, descricao, tipo, categoria, etc.)
        - Analysis results (veredicto, confianca, evidencias, etc.)

        Args:
            basename: Base filename

        Returns:
            Path to CSV file
        """
        filepath = self.output_dir / f"{basename}_analysis.csv"

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            fieldnames = [
                'id', 'descricao', 'tipo', 'categoria', 'prioridade',
                'veredicto', 'confianca', 'evidencias_count', 'fontes',
                'reasoning', 'recomendacoes'
            ]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for req, analysis in zip(self.report.requirements, self.report.analysis_results):
                writer.writerow({
                    'id': req.get('id', ''),
                    'descricao': req.get('descricao', ''),
                    'tipo': req.get('tipo', ''),
                    'categoria': req.get('categoria', ''),
                    'prioridade': req.get('prioridade', ''),
                    'veredicto': analysis.conformity.value,
                    'confianca': f"{analysis.confidence:.2f}",
                    'evidencias_count': len(analysis.evidence),
                    'fontes': ', '.join(analysis.sources),
                    'reasoning': analysis.reasoning,
                    'recomendacoes': '; '.join(analysis.recommendations)
                })

        return filepath

    def to_markdown(self, basename: str) -> Path:
        """
        Export as Markdown report

        Creates a human-readable Markdown document with:
        - Executive summary
        - Detailed requirement-by-requirement analysis
        - Recommendations

        Args:
            basename: Base filename

        Returns:
            Path to Markdown file
        """
        filepath = self.output_dir / f"{basename}_report.md"

        summary = self.report.get_summary()

        md = f"""# Relatório de Análise de Conformidade

**Edital:** {self.report.edital_metadata.get('numero_edital', 'N/A')}
**Órgão:** {self.report.edital_metadata.get('orgao', 'N/A')}
**Modalidade:** {self.report.edital_metadata.get('modalidade', 'N/A')}
**Data da Análise:** {self.report.timestamp}

---

## 📊 Resumo Executivo

| Métrica | Valor |
|---------|-------|
| Total de Requisitos | {summary['total_requirements']} |
| ✅ Conformes | {summary['conforme']} ({summary['conforme_pct']:.1f}%) |
| ❌ Não Conformes | {summary['nao_conforme']} ({summary['nao_conforme_pct']:.1f}%) |
| ⚠️ Revisão Necessária | {summary['revisao']} ({summary['revisao_pct']:.1f}%) |
| 📈 Taxa de Conformidade | {summary['overall_compliance_rate']:.1f}% |

### ⏱️ Performance

| Etapa | Tempo |
|-------|-------|
| Extração de Requisitos | {self.report.pipeline_stats.get('extraction_time', 0):.1f}s |
| Análise de Conformidade | {self.report.pipeline_stats.get('analysis_time', 0):.1f}s |
| Geração de Relatório | {self.report.pipeline_stats.get('report_time', 0):.1f}s |
| **Total** | **{self.report.pipeline_stats.get('total_duration', 0):.1f}s** |

---

## 🚨 Questões Críticas

"""
        critical = self.report.get_critical_issues()
        if critical:
            for item in critical:
                req = item['requirement']
                analysis = item['analysis']
                md += f"""### ❌ {req.get('id', 'N/A')}: {req.get('descricao', 'N/A')}

- **Tipo:** {req.get('tipo', 'N/A')}
- **Categoria:** {req.get('categoria', 'N/A')}
- **Confiança:** {analysis['confidence']:.0%}

**Raciocínio:** {analysis['reasoning']}

---

"""
        else:
            md += "_Nenhuma questão crítica identificada._\n\n---\n\n"

        md += """## ⚠️ Requisitos Necessitando Revisão

"""
        review_needed = self.report.get_review_needed()
        if review_needed:
            for item in review_needed:
                req = item['requirement']
                analysis = item['analysis']
                md += f"""### ⚠️ {req.get('id', 'N/A')}: {req.get('descricao', 'N/A')}

- **Tipo:** {req.get('tipo', 'N/A')}
- **Categoria:** {req.get('categoria', 'N/A')}
- **Confiança:** {analysis['confidence']:.0%}

**Raciocínio:** {analysis['reasoning']}

**Recomendações:**
{chr(10).join(f"- {rec}" for rec in analysis['recommendations'])}

---

"""
        else:
            md += "_Nenhum requisito necessita revisão manual._\n\n---\n\n"

        md += """## 📋 Análise Detalhada

"""

        for req, analysis in zip(self.report.requirements, self.report.analysis_results):
            verdict_emoji = {
                'CONFORME': '✅',
                'NAO_CONFORME': '❌',
                'REVISAO': '⚠️'
            }.get(analysis.conformity.value, '❓')

            md += f"""### {verdict_emoji} {req.get('id', 'N/A')}: {req.get('descricao', 'N/A')}

- **Tipo:** {req.get('tipo', 'N/A')}
- **Categoria:** {req.get('categoria', 'N/A')}
- **Prioridade:** {req.get('prioridade', 'N/A')}
- **Veredicto:** {analysis.conformity.value}
- **Confiança:** {analysis.confidence:.0%}
- **Evidências:** {len(analysis.evidence)} fonte(s): {', '.join(analysis.sources)}

**Raciocínio:** {analysis.reasoning}

**Recomendações:**
{chr(10).join(f"- {rec}" for rec in analysis.recommendations)}

---

"""

        filepath.write_text(md, encoding='utf-8')
        return filepath

    def to_excel(self, basename: str) -> Path:
        """
        Export as Excel (requires openpyxl)

        Creates a multi-sheet workbook with:
        - Sheet 1: Summary statistics
        - Sheet 2: Detailed analysis (one row per requirement)

        Args:
            basename: Base filename

        Returns:
            Path to Excel file

        Raises:
            ImportError: If openpyxl is not installed
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError(
                "openpyxl required for Excel export. "
                "Install with: pip install openpyxl"
            )

        filepath = self.output_dir / f"{basename}_analysis.xlsx"
        wb = openpyxl.Workbook()

        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Resumo"
        summary = self.report.get_summary()

        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        ws_summary['A1'] = "Métrica"
        ws_summary['B1'] = "Valor"
        ws_summary['A1'].fill = header_fill
        ws_summary['B1'].fill = header_fill
        ws_summary['A1'].font = header_font
        ws_summary['B1'].font = header_font

        # Data
        rows = [
            ("Total de Requisitos", summary['total_requirements']),
            ("Conformes", summary['conforme']),
            ("Não Conformes", summary['nao_conforme']),
            ("Revisão Necessária", summary['revisao']),
            ("Taxa de Conformidade (%)", f"{summary['overall_compliance_rate']:.1f}%")
        ]

        for idx, (metric, value) in enumerate(rows, start=2):
            ws_summary[f'A{idx}'] = metric
            ws_summary[f'B{idx}'] = value

        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 15

        # Sheet 2: Detailed Analysis
        ws_detail = wb.create_sheet("Análise Detalhada")
        headers = [
            'ID', 'Descrição', 'Tipo', 'Categoria', 'Prioridade',
            'Veredicto', 'Confiança', 'Evidências', 'Fontes', 'Raciocínio'
        ]
        ws_detail.append(headers)

        # Style header row
        for cell in ws_detail[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add data rows
        for req, analysis in zip(self.report.requirements, self.report.analysis_results):
            ws_detail.append([
                req.get('id', ''),
                req.get('descricao', ''),
                req.get('tipo', ''),
                req.get('categoria', ''),
                req.get('prioridade', ''),
                analysis.conformity.value,
                analysis.confidence,
                len(analysis.evidence),
                ', '.join(analysis.sources),
                analysis.reasoning
            ])

        # Adjust column widths
        ws_detail.column_dimensions['A'].width = 12
        ws_detail.column_dimensions['B'].width = 50
        ws_detail.column_dimensions['C'].width = 12
        ws_detail.column_dimensions['D'].width = 15
        ws_detail.column_dimensions['E'].width = 12
        ws_detail.column_dimensions['F'].width = 15
        ws_detail.column_dimensions['G'].width = 12
        ws_detail.column_dimensions['H'].width = 12
        ws_detail.column_dimensions['I'].width = 30
        ws_detail.column_dimensions['J'].width = 60

        wb.save(filepath)
        return filepath
